from datetime import datetime
from PyQt5.sip import delete
import openpyxl
import os

class attendance:
    def __init__(self):
        self.day = ['월', '화', '수', '목', '금', '토', '일']
        self.path_data = os.path.join(os.getcwd(), 'data.xlsx')
        self.path_schedule = os.path.join(os.getcwd(), 'schedule.xlsx') 
        self.path_log = os.path.join(os.getcwd(),'log' ,str(datetime.today().strftime("%Y-%m-%d")) + f"-{self.day[datetime.today().weekday()]}.txt")
        self.col_num_schedule = 0
        self.user_names = list()
        self.attend_user_names = list()
        self.log = ""
        self.data = self.data_load()

        try:
            self.schedule_set_data()

        except FileNotFoundError:
            self.schedule_init()
            self.schedule_set_data()

        try:
            self.log = self.log_load()

        except FileNotFoundError:
            self.log_init()

    def log_init(self):
        with open(self.path_log, "w") as f:
            pass

    def log_load(self):
        with open(self.path_log, "r") as f:
            logs = f.read()
        return logs

    def log_save(self, log):
        with open(self.path_log, "w") as f:
            f.write(log)

    def data_load(self):
        data = dict()
        row = 2
        data_files = openpyxl.load_workbook(self.path_data)
        data_xlsx = data_files.active      

        while True:
            if data_xlsx.cell(row=row, column=1).value != None:
                user_id = row - 2
                user = str(data_xlsx.cell(row=row, column=1).value)
                first_day = str(data_xlsx.cell(row=row, column=2).value).split(',')
                second_day = (lambda day: [] if day == None else str(day).split(','))(data_xlsx.cell(row=row, column=3).value)
                early_leave_count = int(data_xlsx.cell(row=row, column=4).value)
                late_count = int(data_xlsx.cell(row=row, column=5).value)
                absent_count = int(data_xlsx.cell(row=row, column=6).value)
                pos = int(data_xlsx.cell(row=row, column=7).value)
                data[user] = {"id" : user_id, "first_day" : first_day, "second_day" : second_day, "early_leave_count" : early_leave_count, "late_count": late_count, "absent_count" : absent_count, "pos" : pos, "first_check_time": None , "second_check_time" : None}
                
                self.user_names.append(user)
                row = row + 1

            elif data_xlsx.cell(row=row, column=1).value == None:
                break
            
        data_files.close()

        return data

    def save_data(self, users):
        data_files = openpyxl.load_workbook(self.path_data)
        data_xlsx = data_files.active
        schedule_files = openpyxl.load_workbook(self.path_schedule)
        sf = schedule_files['1교시']

        for user in users:
            row = self.data[user]["id"] + 2
            data_xlsx.cell(row=row, column=4).value = sf.cell(row=row, column=5).value
            data_xlsx.cell(row=row, column=5).value = sf.cell(row=row, column=4).value
            data_xlsx.cell(row=row, column=6).value = sf.cell(row=row, column=2).value
        
        data_files.save(self.path_data)


    def schedule_init(self):
        data_files = openpyxl.load_workbook(self.path_data)
        df = data_files.active

        schedule_files = openpyxl.Workbook()
        sf_c1 = schedule_files.active
        sf_c2 = schedule_files.create_sheet()
        sf_c1.title = "1교시"
        sf_c1.cell(row=1, column=1).value = "이름"
        sf_c1.cell(row=1, column=2).value = "총 결석 수"
        sf_c1.cell(row=1, column=3).value = "결석 수"
        sf_c1.cell(row=1, column=4).value = "지각 수"
        sf_c1.cell(row=1, column=5).value = "조퇴 수"

        sf_c2.title = "2교시"
        sf_c2.cell(row=1, column=1).value = "이름"
        sf_c2.cell(row=1, column=2).value = "총 결석 수"
        sf_c2.cell(row=1, column=3).value = "결석 수"
        sf_c2.cell(row=1, column=4).value = "지각 수"
        sf_c2.cell(row=1, column=5).value = "조퇴 수"

        for user in self.user_names:
            row = self.data[user]["id"] + 2

            sf_c1.cell(row=row, column=1).value = user
            sf_c1.cell(row=row, column=2).value = self.sum_absent_count(self.data[user]["absent_count"], self.data[user]["late_count"], self.data[user]["early_leave_count"])
            sf_c1.cell(row=row, column=3).value = self.data[user]["absent_count"]
            sf_c1.cell(row=row, column=4).value = self.data[user]["late_count"]
            sf_c1.cell(row=row, column=5).value = self.data[user]["early_leave_count"]

            sf_c2.cell(row=row, column=1).value = user
            sf_c2.cell(row=row, column=2).value = self.sum_absent_count(self.data[user]["absent_count"], self.data[user]["late_count"], self.data[user]["early_leave_count"])
            sf_c2.cell(row=row, column=3).value = self.data[user]["absent_count"]
            sf_c2.cell(row=row, column=4).value = self.data[user]["late_count"]
            sf_c2.cell(row=row, column=5).value = self.data[user]["early_leave_count"]

        schedule_files.save(self.path_schedule)

    def schedule_set_data(self):
        schedule_files = openpyxl.load_workbook(self.path_schedule)
        sf_c1 = schedule_files['1교시']
        sf_c2 = schedule_files['2교시']
        today_date = str(datetime.today().strftime("%Y-%m-%d")) + f" {self.day[datetime.today().weekday()]}"
        print(today_date)
        self.col_num_schedule = 6
        while True: 
            date = sf_c1.cell(row=1, column=self.col_num_schedule).value
            if date == None and date != today_date:
                print(1)
                sf_c1.cell(row=1, column=self.col_num_schedule).value = today_date
                sf_c2.cell(row=1, column=self.col_num_schedule).value = today_date
                break
            elif date == today_date:
                print(2)
                break
            else:
                print(3)
                self.col_num_schedule = self.col_num_schedule + 1

        row = 2
        while True:
            user = sf_c1.cell(row=row, column=1).value
            first_check_time = sf_c1.cell(row=row, column=self.col_num_schedule).value
            second_check_time = sf_c2.cell(row=row, column=self.col_num_schedule).value
            if user == None:
                break
            self.data[user]["first_check_time"] = first_check_time
            self.data[user]["second_check_time"] = second_check_time
            row = row + 1

        schedule_files.save(self.path_schedule)

    def save_schedule(self, users):
        schedule_files = openpyxl.load_workbook(self.path_schedule)
        sf_c1 = schedule_files['1교시']
        sf_c2 = schedule_files['2교시']

        for user in users:
            row = self.data[user]["id"] + 2
             
            # sf_c1.cell(row=row, column=2).value = self.sum_absent_count(self.data[user]["absent_count"], self.data[user]["late_count"], self.data[user]["early_leave_count"])
            # sf_c1.cell(row=row, column=3).value = self.data[user]["absent_count"]
            # sf_c1.cell(row=row, column=4).value = self.data[user]["late_count"]
            # sf_c1.cell(row=row, column=5).value = self.data[user]["early_leave_count"]
            sf_c1.cell(row=row, column=self.col_num_schedule).value = self.data[user]["first_check_time"]

            # sf_c2.cell(row=row, column=2).value = self.sum_absent_count(self.data[user]["absent_count"], self.data[user]["late_count"], self.data[user]["early_leave_count"])
            # sf_c2.cell(row=row, column=3).value = self.data[user]["absent_count"]
            # sf_c2.cell(row=row, column=4).value = self.data[user]["late_count"]
            # sf_c2.cell(row=row, column=5).value = self.data[user]["early_leave_count"]
            sf_c2.cell(row=row, column=self.col_num_schedule).value = self.data[user]["second_check_time"]      
            print(user, self.data[user]["first_check_time"] ,self.data[user]["second_check_time"])
        schedule_files.save(self.path_schedule)     

    def attend_check(self, user, class_time):
        if class_time == 1:
            self.data[user]["first_check_time"] = "출석"

        elif class_time == 2 and self.data[user]["second_day"] != []:    
            self.data[user]["second_check_time"] = "출석"


    def absent_check(self, user, class_time):
        # print(f"{class_time} jk")
        if class_time == 1:
            self.data[user]["first_check_time"] = "결석"

        elif class_time == 2 and self.data[user]["second_day"] != []:    
            self.data[user]["second_check_time"] = "결석"

            

    def late_check(self, user, class_time):
        if class_time == 1:
            self.data[user]["first_check_time"] = "지각"

        elif class_time == 2 and self.data[user]["second_day"] != []:    
            self.data[user]["second_check_time"] = "지각"


    def early_leave_check(self, user, class_time):
        if class_time == 1:
            self.data[user]["first_check_time"] = "조퇴"

        elif class_time == 2 and self.data[user]["second_day"] != []:    
            self.data[user]["second_check_time"] = "조퇴"

    def schedule_count(self):
        schedule_files = openpyxl.load_workbook(self.path_schedule)
        sf_c1 = schedule_files['1교시']
        sf_c2 = schedule_files['2교시']

        for user in self.user_names:
            row = self.data[user]["id"] + 2
            late = 0
            absent = 0
            early = 0

            for col in range(6, self.col_num_schedule + 1):
                first = sf_c1.cell(row=row, column=col).value
                second = sf_c2.cell(row=row, column=col).value

                if first == "출석":
                    if second == "출석" or second == None:
                        pass
                        
                    elif second == "지각":
                        late = late + 1
                    
                    elif second == "결석" or second == "조퇴":
                        early = early + 1

                elif first == "지각":
                    late = late + 1

                    if second == "출석" or second == None:
                        pass
                        
                    elif second == "지각":
                        late = late + 1
                    
                    elif second == "결석" or second == "조퇴":
                        early = early + 1

                elif first == "조퇴":
                    early = early + 1
                
                elif first == "결석":
                    absent = absent + 1


                if first == None and second == "출석":
                    pass

                elif first == None and second == "지각":
                    late = late + 1

                elif first == None and second == "조퇴":
                    early = early + 1
                
                elif first == None and second == "결석":
                    absent = absent + 1





            sf_c1.cell(row=row, column=2).value = self.sum_absent_count(absent, late, early)
            sf_c1.cell(row=row, column=3).value = absent
            sf_c1.cell(row=row, column=4).value = late
            sf_c1.cell(row=row, column=5).value = early

            sf_c2.cell(row=row, column=2).value = self.sum_absent_count(absent, late, early)
            sf_c2.cell(row=row, column=3).value = absent
            sf_c2.cell(row=row, column=4).value = late
            sf_c2.cell(row=row, column=5).value = early

        schedule_files.save(self.path_schedule)


    def sum_absent_count(self, absent_count, late_count, early_leave_count):
        sum_absent = absent_count + (late_count + early_leave_count) // 3   
        return sum_absent

    def delete_check(self):
        data_files = openpyxl.load_workbook(self.path_data)
        data_xlsx = data_files.active     
        
        row = 2
        delete_user = list()
        st = 0

        while True:
            if data_xlsx.cell(row=row, column=1).value != None:
                if data_xlsx.cell(row=row, column=6).value >= 2:
                    user = str(data_xlsx.cell(row=row, column=1).value)
                    delete_user.append(user)
                    st = 1
                row = row + 1

            elif data_xlsx.cell(row=row, column=1).value == None:
                break

        if st == 1:
            return delete_user
        else:
            return 1

    def delete_user(self, user):
        data_files = openpyxl.load_workbook(self.path_data)
        data_xlsx = data_files.active   
        schedule_files = openpyxl.load_workbook(self.path_schedule)
        sf_c1 = schedule_files['1교시']
        sf_c2 = schedule_files['2교시']

        col = self.data[user]['id'] + 2

        print(col)

        
# att = attendance()

# att.save_schedule()
# print(att.data)
