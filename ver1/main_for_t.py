from datetime import datetime
import openpyxl
import os

class attendance:
    def __init__(self):
        self.day = ['월', '화', '수', '목', '금', '토', '일']
        self.path_data = os.path.join(os.getcwd(), 'data.xlsx')
        self.path_schedule = os.path.join(os.getcwd(), 'schedule.xlsx') 
        self.col_date = 0
        self.data = self.load()
        self.users = self.user_names()
        try:
            self.schedule_check_init()
        
        except FileNotFoundError:
            self.schedule_init()
            self.schedule_check_init()

    def user_names(self):
        users = list()
        num = 2
        files = openpyxl.load_workbook(self.path_data)
        xlsx = files.active
        while True:
            if xlsx.cell(row=num, column=1).value != None:
                users.append(str(xlsx.cell(row=num, column=1).value))
                num = num + 1
            else:
                break
        return users

    def load(self):
        data = dict()
        num = 2
        files = openpyxl.load_workbook(self.path_data)
        xlsx = files.active

        while True:
            if xlsx.cell(row=num, column=1).value != None:
                data[str(xlsx.cell(row=num, column=1).value)] = {"last_date" : str(xlsx.cell(row=num, column=2).value).split(' ')[0], "count" : int(xlsx.cell(row=num, column=5).value), "first_day" : str(xlsx.cell(row=num, column=3).value).split(','),"second_day" : (lambda text: None if text == None else str(xlsx.cell(row=num, column=4).value).split(','))(str(xlsx.cell(row=num, column=3).value)) , "check_time": None, "pos":str(xlsx.cell(row=num, column=6).value)}
            elif xlsx.cell(row=num, column=1).value == None:
                break
            num = num + 1
        # print(str(xlsx.cell(row=1, column=2).value).split(' '))
        # print(data["031423-이건우"]["last_date"])
        return data

    def save(self):
        files1 = openpyxl.load_workbook(self.path_data)
        ds = files1.active

        files2 = openpyxl.load_workbook(self.path_schedule, read_only=False, data_only=True)
        ws = files2.active

        num = 2
        while True:
            if ds.cell(row=num, column=1).value != None:
                ds.cell(row=num, column=2).value = self.data[ds.cell(row=num, column=1).value]["last_date"]
                ds.cell(row=num, column=5).value = self.data[ds.cell(row=num, column=1).value]["count"]    
                ws.cell(row=num,column=2).value = self.data[ds.cell(row=num, column=1).value]["count"]   
                
            elif ds.cell(row=num, column=1).value == None:
                break
            num = num + 1
        
        files1.save(self.path_data)
        files2.save(self.path_schedule)

    def schedule_init(self):
        files = openpyxl.load_workbook(self.path_data)
        ds = files.active

        wb = openpyxl.Workbook()
        ws = wb.active

        num = 2
        ws.cell(row=1, column=1).value = "학번-이름"
        ws.cell(row=1, column=2).value = "출석 수"

        while True:
            if ds.cell(row=num, column=1).value != None:
                ws.cell(row=num, column=1).value = str(ds.cell(row=num, column=1).value)
                ws.cell(row=num, column=2).value = self.data[str(ds.cell(row=num, column=1).value)]["count"]    
                
            elif ds.cell(row=num, column=1).value == None:
                break
            num = num + 1
        
        wb.save(self.path_schedule)


    def schedule_check_init(self):
        files = openpyxl.load_workbook(self.path_schedule, read_only=False, data_only=True)
        ds = files.active

        self.col_date = 3
        while True:
            if ds.cell(row=1, column=self.col_date).value == None and ds.cell(row=1, column=self.col_date).value != str(datetime.today().strftime("%Y-%m-%d")):
                ds.cell(row=1, column=self.col_date).value = str(datetime.today().strftime("%Y-%m-%d")) + f" {self.day[datetime.today().weekday()]}"
                break

            elif ds.cell(row=1, column=self.col_date).value == str(datetime.today().strftime("%Y-%m-%d")) + f" {self.day[datetime.today().weekday()]}":
                num = 2
                while True:
                    if ds.cell(row=num, column=1).value != None:
                       self.data[str(ds.cell(row=num, column=1).value)]["check_time"] = ds.cell(row=num, column=self.col_date).value
                    #    print(ds.cell(row=num, column=self.col_date).value)
                    else:
                        break
                    num = num + 1

                break

            else:
                pass
            
            self.col_date = self.col_date + 1
                
        files.save(self.path_schedule)

    def schedule_check(self, user):
        files = openpyxl.load_workbook(self.path_schedule, read_only=False, data_only=True)
        ds = files.active

        user_num = 2
        while True:
            if str(ds.cell(row=user_num, column=1).value) == str(user):
                ds.cell(row=user_num, column=self.col_date).value = str(self.data[str(user)]["check_time"])
                ds.cell(row=user_num, column=2).value = self.data[str(user)]["count"]    
                print(user_num)
                break
            user_num = user_num + 1
        files.save(self.path_schedule)
    

# at = attendance()
# print(at.data["031401-강일우"])

# at.run()
# at.schedule_init()
# at.schedule_edit()
