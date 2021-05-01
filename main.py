from datetime import datetime
import openpyxl

class attendance:
    def __init__(self):
        self.day = ['월', '화', '수', '목', '금', '토', '일']
        self.col_date = 0
        self.data = self.load()
        self.users = self.user_names()
        try:
            self.schedule_check_init()
        
        except FileNotFoundError:
            self.schedule_init()
            self.schedule_check_init()

    def user_names(self):
        users = list()
        num = 1
        files = openpyxl.load_workbook('data.xlsx')
        xlsx = files.active
        while True:
            if xlsx.cell(row=num, column=1).value != None:
                users.append(str(xlsx.cell(row=num, column=1).value))
                num = num + 1
            else:
                break
        return users

    def load(self):
        data = dict()
        num = 1
        files = openpyxl.load_workbook('data.xlsx')
        xlsx = files.active

        while True:
            if xlsx.cell(row=num, column=1).value != None:
                data[str(xlsx.cell(row=num, column=1).value)] = {"last_date" : str(xlsx.cell(row=num, column=2).value).split(' ')[0], "count" : int(xlsx.cell(row=num, column=5).value), "first_day" : str(xlsx.cell(row=num, column=3).value).split(','),"second_day" : (lambda text: None if text == None else str(xlsx.cell(row=num, column=4).value).split(','))(str(xlsx.cell(row=num, column=3).value)) , "check_time": None, "pos":str(xlsx.cell(row=num, column=6).value)}
            elif xlsx.cell(row=num, column=1).value == None:
                break
            num = num + 1
        # print(str(xlsx.cell(row=1, column=2).value).split(' '))
        # print(data["031423-이건우"]["last_date"])
        return data

    def save(self):
        files1 = openpyxl.load_workbook('data.xlsx')
        ds = files1.active

        files2 = openpyxl.load_workbook('schedule.xlsx', read_only=False, data_only=True)
        ws = files2.active

        num = 1
        while True:
            if ds.cell(row=num, column=1).value != None:
                ds.cell(row=num, column=2).value = self.data[ds.cell(row=num, column=1).value]["last_date"]
                ds.cell(row=num, column=5).value = self.data[ds.cell(row=num, column=1).value]["count"]    
                ws.cell(row=num+1,column=2).value = self.data[ds.cell(row=num, column=1).value]["count"]   
                
            elif ds.cell(row=num, column=1).value == None:
                break
            num = num + 1
        
        files1.save("data.xlsx")
        files2.save('schedule.xlsx')
    
    def schedule_init(self):
        files = openpyxl.load_workbook('data.xlsx')
        ds = files.active

        wb = openpyxl.Workbook()
        ws = wb.active

        num = 2
        ws.cell(row=1, column=1).value = "학번-이름"
        ws.cell(row=1, column=2).value = "출석 수"

        while True:
            if ds.cell(row=num-1, column=1).value != None:
                ws.cell(row=num, column=1).value = str(ds.cell(row=num-1, column=1).value)
                ws.cell(row=num, column=2).value = self.data[str(ds.cell(row=num-1, column=1).value)]["count"]    
                
            elif ds.cell(row=num-1, column=1).value == None:
                break
            num = num + 1
        
        wb.save("schedule.xlsx")


    def schedule_check_init(self):
        files = openpyxl.load_workbook('schedule.xlsx', read_only=False, data_only=True)
        ds = files.active

        self.col_date = 3
        while True:
            if ds.cell(row=1, column=self.col_date).value == None and ds.cell(row=1, column=self.col_date).value != str(datetime.today().strftime("%Y-%m-%d")):
                ds.cell(row=1, column=self.col_date).value = str(datetime.today().strftime("%Y-%m-%d")) + f" {self.day[datetime.today().weekday()]}"
                break

            elif ds.cell(row=1, column=self.col_date).value == str(datetime.today().strftime("%Y-%m-%d")) + f" {self.day[datetime.today().weekday()]}":
                num = 2
                while True:
                    if ds.cell(row=num, column=1).value != None:
                       self.data[str(ds.cell(row=num, column=1).value)]["check_time"] = ds.cell(row=num, column=self.col_date).value
                    #    print(ds.cell(row=num, column=self.col_date).value)
                    else:
                        break
                    num = num + 1

                break

            else:
                pass
            
            self.col_date = self.col_date + 1
                
        files.save("schedule.xlsx")

    def schedule_check(self, user):
        files = openpyxl.load_workbook('schedule.xlsx', read_only=False, data_only=True)
        ds = files.active

        user_num = 2
        while True:
            if str(ds.cell(row=user_num, column=1).value) == str(user):
                if ds.cell(row=user_num, column=self.col_date).value == None:
                    ds.cell(row=user_num, column=self.col_date).value = str(self.data[str(user)]["check_time"]) + "/"
                else:
                    ds.cell(row=user_num, column=self.col_date).value = str(ds.cell(row=user_num, column=self.col_date).value) + str(self.data[str(user)]["check_time"]) + "/"
                ds.cell(row=user_num, column=2).value = self.data[str(user)]["count"]    
                break
            user_num = user_num + 1
        files.save("schedule.xlsx")
    
    def check_time(self, time):
        if time == None:
            return True
            
        time = time.split('/')[0]
        time1 = datetime(datetime.now().year, datetime.now().month, datetime.now().day, int(time.split(':')[0]), int(time.split(':')[1]), int(time.split(':')[2]))
        if (datetime.now() - time1).seconds / 3600 >= 1:
            return True
        
        else:
            return False

    # print(data["031402-강준"]["count"])
    def run(self):
        while True:
            print("야자 출석 명단")
            # print(len(str(self.data["031401-강일우"]["check_time"])))
            # print(str(self.data["031401-강일우"]["check_time"]))
            # print(len(str(self.data["031408-김서현"]["check_time"])))
            for i in self.users:
                if self.day[datetime.today().weekday()] in self.data[i]["first_day"] and self.data[i]["last_date"] != str(datetime.today().strftime("%Y-%m-%d")):
                    if self.day[datetime.today().weekday()] in self.data[i]["second_day"]:
                        print(f'{i} : X | X')
                    else:
                        print(f'{i} : X')
                elif self.day[datetime.today().weekday()] in self.data[i]["first_day"] and self.data[i]["last_date"] == str(datetime.today().strftime("%Y-%m-%d")):
                    if self.day[datetime.today().weekday()] in self.data[i]["second_day"]:
                        if self.data[i]["check_time"] == None or len(str(self.data[i]["check_time"])) <= 10:
                            print(f'{i} : O | X')
                        elif len(str(self.data[i]["check_time"])) > 10 :
                            print(f'{i} : O | O')
                    else:
                        print(f'{i} : O')
            print('\n',end='')

            print("학번-이름 : ",end='')
            user = input()

            if user == "stop":
                break
            
            if user == "save":
                self.save()
                break

            try:
                if str(datetime.today().strftime("%Y-%m-%d")) != self.data[user]["last_date"] and self.data[user]["check_time"] == None:
                    self.data[user]["last_date"] = str(datetime.today().strftime("%Y-%m-%d"))
                    self.data[user]["check_time"] = str(datetime.today().strftime("%H:%M:%S")) 
                    self.data[user]["count"] = self.data[user]["count"] + 1
                    self.schedule_check(user)
                    print("출석완료.")
                    print('\n',end='')

                elif self.check_time(self.data[user]["check_time"]):
                    self.data[user]["check_time"] = self.data[user]["check_time"] + str(datetime.today().strftime("%H:%M:%S"))
                    self.schedule_check(user)
                    print("출석완료.")
                    print('\n',end='')
                
                elif str(datetime.today().strftime("%Y-%m-%d")) == self.data[user]["last_date"] and self.check_time(self.data[user]["check_time"]) == False:
                    print("이미 춣석 하셨습니다.")
                    print('\n',end='')
                

            except KeyError:
                pass
    

# at = attendance()
# print(at.data["031401-강일우"])

# at.run()
# at.schedule_init()
# at.schedule_edit()
